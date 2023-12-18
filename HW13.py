import openpyxl
from openpyxl.styles import Font, Border, Side
import os
file_paths=["HW13 (1).xlsx", "HW13 (2).xlsx", "HW13 (3).xlsx"]
sorted_inf=[]
for file_path in file_paths:
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook.active
    inf=[]
    for stroki in sheet.iter_rows(min_row=1, values_only=True):
        inf.append(stroki)
    sorted_inf.extend(inf)
print(sorted_inf)
sorted_inf = sorted(sorted_inf, key=lambda x: x[0], reverse=True)
output_file='HW13.xlsx'
if os.path.exists(output_file):
    workbook = openpyxl.load_workbook(output_file)
else:
    workbook = openpyxl.Workbook()
sheet=workbook.active
for strok_num, strok_data in enumerate(sorted_inf, start=1):
    for col_num, value in enumerate(strok_data, start=1):
        sheet.cell(row=strok_num, column=col_num, value=value).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
workbook.save(output_file)